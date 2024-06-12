import unittest
from flatten_json import flatten_json, FlattenJSONError
import csv
import openpyxl

class TestFlattenJSON(unittest.TestCase):
    def test_nested_object(self):
        data = {"a": {"b": {"c": 1}}}
        result = flatten_json(data)
        self.assertEqual(result, {"a.b.c": 1})

    def test_array(self):
        data = {"a": [1, 2, 3]}
        result = flatten_json(data)
        self.assertEqual(result, {"a.0": 1, "a.1": 2, "a.2": 3})

    def test_array_concatenate(self):
        data = {"a": [1, 2, 3]}
        result = flatten_json(data, array_handling="concatenate")
        self.assertEqual(result, {"a": "1, 2, 3"})

    def test_array_ignore(self):
        data = {"a": [1, 2, 3]}
        result = flatten_json(data, array_handling="ignore")
        self.assertEqual(result, {})

    def test_mixed_types(self):
        data = {"a": {"b": [1, "two", True, None]}}
        result = flatten_json(data)
        self.assertEqual(result, {"a.b.0": 1, "a.b.1": "two", "a.b.2": True, "a.b.3": None})

    def test_key_collision_suffix(self):
        data = {"a": {"b": 1}, "a.b": 2}
        result = flatten_json(data, handle_collision="suffix")
        self.assertEqual(result, {"a.b": 1, "a.b_duplicate": 2})

    def test_key_collision_counter(self):
        data = {"a": {"b": 1}, "a.b": 2}
        result = flatten_json(data, handle_collision="counter")
        self.assertEqual(result, {"a.b": 1, "a.b_1": 2})

    def test_key_collision_error(self):
        data = {"a": {"b": 1}, "a.b": 2}
        with self.assertRaises(KeyError):
            flatten_json(data, handle_collision="error")

    def test_custom_handlers(self):
        data = {"a": {"b": 1.2345}}
        custom_handlers = {float: lambda x: round(x, 2)}
        result = flatten_json(data, custom_handlers=custom_handlers)
        self.assertEqual(result, {"a.b": 1.23})

    def test_invalid_input(self):
        with self.assertRaises(FlattenJSONError):
            flatten_json(123)  # Invalid input type (not dict or list)

    def test_value_transform(self):
        data = {"a": {"b": "2024-06-12"}}
        value_transform = lambda x: x.replace("-", "/") if isinstance(x, str) else x
        result = flatten_json(data, value_transform=value_transform)
        self.assertEqual(result, {"a.b": "2024/06/12"})

    def test_key_filter(self):
        data = {"a": {"b": 1, "c": 2}}
        result = flatten_json(data, key_filter=["a.b"])
        self.assertEqual(result, {"a.b": 1})

    def test_max_depth(self):
        data = {"a": {"b": {"c": 1}}}
        result = flatten_json(data, max_depth=1)
        self.assertEqual(result, {"a.b": {"c": 1}})

    def test_sort_keys(self):
        data = {"b": 1, "a": 2}
        result = flatten_json(data, sort_keys=True)
        self.assertEqual(result, {"a": 2, "b": 1})

    def test_parallel_processing(self):
        data = {"a": [{"b": i} for i in range(1000)]}
        result = flatten_json(data, parallel=True)
        expected = {f"a.{i}.b": i for i in range(1000)}
        self.assertEqual(result, expected)

    def test_output_csv(self):
        data = {"a": 1, "b": 2}
        flatten_json(data, output_format="csv")
        with open("flattened_data.csv", "r") as file:
            reader = csv.reader(file)
            output = {rows[0]: int(rows[1]) for rows in reader}
        self.assertEqual(output, data)

    def test_output_xlsx(self):
        data = {"a": 1, "b": 2}
        flatten_json(data, output_format="xlsx")
        workbook = openpyxl.load_workbook("flattened_data.xlsx")
        sheet = workbook.active
        output = {sheet.cell(row=i, column=1).value: sheet.cell(row=i, column=2).value for i in range(1, sheet.max_row + 1)}
        self.assertEqual(output, data)

if __name__ == "__main__":
    unittest.main(argv=[''], exit=False)
